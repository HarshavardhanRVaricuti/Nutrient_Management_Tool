import pandas as pd
import os
from openpyxl import load_workbook
from oletools.olevba import VBA_Parser 


def extract_excel_formulas(file_path, sheet_name):
    wb = load_workbook(filename=file_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
        return
    sheet = wb[sheet_name]
    
    for row in range(1, sheet.max_row + 1):
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f"Cell {cell.coordinate}: {cell.value}")

def extract_vba_macros(file_path):
    vba_parser = VBA_Parser(file_path)
    if vba_parser.detect_vba_macros():

        for (filename, stream_path, vba_filename, vba_code) in vba_parser.extract_macros():
            print(f"Macro found in {vba_filename}")
            print("-" * 40)
            print(vba_code)
            print("-" * 40)
        vba_parser.close()
    else:
        print("No VBA macros found.")

if __name__ == "__main__":

    file_path = os.path.join(os.getcwd(), 'data', 'tamu_legacy_nutrient_tool.xlsm')    
    if os.path.exists(file_path):
        extract_excel_formulas(file_path, sheet_name = "Calculator")
        extract_vba_macros(file_path)
    else:
        print("File not found.")